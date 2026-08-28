"""
generate_tables_pdf.py
-----------------------
Erzeugt aus einer "game_data.xlsx"-Datei ein PDF mit vier Tabellen:

1. "Board BaseValues"            -> Wert-Tabelle mit rot umrandetem Startwert
2. "Board Heiztabelle SP/Budget" -> kombinierte Heiztabelle (SP / Budget je Zelle)
3. "Board WP Netzbezug"          -> Effizienz-Tabelle (Effizienz1-5, Effizienz0 wird ignoriert)
4. "Netzbezug Impact"            -> Netzbezug / Budget / SP Runde 1-4

Das Skript sucht die benötigten Spalten über die Kopfzeile (Zeile 1) der
jeweiligen Sheets, damit es auch dann funktioniert, wenn sich Spalten-
Reihenfolge oder -Anzahl in einer neuen Version der Datei ändert.

Aufruf:
    python generate_tables_pdf.py [input.xlsx] [output.pdf]

Standardwerte: input.xlsx = "game_data.xlsx", output.pdf = "game_data_tables.pdf"
"""

import sys
import re
import openpyxl

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)

# ----------------------------------------------------------------------
# Hilfsfunktionen
# ----------------------------------------------------------------------

def get_header_map(ws, header_row=1):
    """Liefert ein Dict {Spaltenname: Spaltenindex (1-basiert)} für eine Kopfzeile."""
    header = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            header[str(cell.value).strip()] = cell.column
    return header


def col_letter_for(header_map, name):
    """Gibt die Spaltennummer für einen exakten Header-Namen zurück (oder None)."""
    return header_map.get(name)


def fmt_num(value):
    """Formatiert Zahlen hübsch: alle Kommazahlen werden auf ganze Zahlen gerundet."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(round(value)))
    return str(value)


# ----------------------------------------------------------------------
# Tabelle 1: Board BaseValues
# ----------------------------------------------------------------------

def build_base_values_table(wb, styles):
    ws = wb["Board BaseValues"]
    header_map = get_header_map(ws)

    wert_col = col_letter_for(header_map, "Wert")
    start_col = col_letter_for(header_map, "Start (0-basiert)")

    # Alle "ValuesN"-Spalten finden und nach N sortieren
    values_cols = []
    for name, col in header_map.items():
        m = re.fullmatch(r"Values(\d+)", name)
        if m:
            values_cols.append((int(m.group(1)), col))
    values_cols.sort(key=lambda x: x[0])

    if wert_col is None or start_col is None or not values_cols:
        return None  # Sheet/Struktur nicht gefunden

    n_values = len(values_cols)
    header_row = ["Wert"] + [str(i) for i, _ in values_cols]
    data = [header_row]

    # merkt sich für jede Datenzeile, welche Zelle (row, col) rot umrandet werden soll
    highlight_cells = []

    row_idx = 2
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        wert_value = row[wert_col - 1].value
        if wert_value is None:
            continue
        start_value = row[start_col - 1].value
        line = [str(wert_value)]
        for pos, (_, col) in enumerate(values_cols):
            line.append(fmt_num(row[col - 1].value))
        data.append(line)

        if isinstance(start_value, (int, float)):
            start_idx = int(start_value)
            if 0 <= start_idx < n_values:
                # Tabellen-Koordinaten: Spalte = start_idx+1 (Wert-Spalte ist 0),
                # Zeile = aktuelle Datenzeile in der `data`-Liste
                highlight_cells.append((start_idx + 1, len(data) - 1))
        row_idx += 1

    if len(data) <= 1:
        return None

    col_widths = [4.5 * cm] + [1.4 * cm] * n_values
    table = Table(data, colWidths=col_widths, repeatRows=1)

    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]
    # rote Umrandung für den Startwert jeder Zeile
    for col, row in highlight_cells:
        style.append(("BOX", (col, row), (col, row), 1.5, colors.red))

    table.setStyle(TableStyle(style))

    elements = [Paragraph("1. Board BaseValues", styles["Heading2"]), Spacer(1, 6), table]
    return elements


# ----------------------------------------------------------------------
# Tabelle 2: Board Heiztabelle SP + Budget kombiniert
# ----------------------------------------------------------------------

def build_heiztabelle_combined(wb, styles):
    if "Board Heiztabelle SP" not in wb.sheetnames or "Board Heiztabelle Budget" not in wb.sheetnames:
        return None

    ws_sp = wb["Board Heiztabelle SP"]
    ws_budget = wb["Board Heiztabelle Budget"]

    header_sp = get_header_map(ws_sp)
    header_budget = get_header_map(ws_budget)

    categories = ["Gas", "Biomasse", "Fernwärme", "Grünes Gas", "Wärmepumpe"]

    ws_col_sp = col_letter_for(header_sp, "WS")
    ws_col_budget = col_letter_for(header_budget, "WS")
    if ws_col_sp is None or ws_col_budget is None:
        return None

    # WS -> Zeilenindex je Sheet
    def ws_row_map(ws, ws_col):
        m = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            v = row[ws_col - 1].value
            if isinstance(v, (int, float)):
                m[int(v)] = row[0].row
        return m

    sp_rows = ws_row_map(ws_sp, ws_col_sp)
    budget_rows = ws_row_map(ws_budget, ws_col_budget)

    all_ws = sorted(set(sp_rows) | set(budget_rows), reverse=True)  # 9 oben, 0 unten

    header_row = ["WS"] + categories
    data = [header_row]

    for ws_value in all_ws:
        line = [str(ws_value)]
        for cat in categories:
            sp_val = None
            budget_val = None
            if cat in header_sp and ws_value in sp_rows:
                sp_val = ws_sp.cell(row=sp_rows[ws_value], column=header_sp[cat]).value
            if cat in header_budget and ws_value in budget_rows:
                budget_val = ws_budget.cell(row=budget_rows[ws_value], column=header_budget[cat]).value
            line.append(f"SP: {fmt_num(sp_val)}\nBudget: {fmt_num(budget_val)}")
        data.append(line)

    # Zellen mit Zeilenumbruch als Paragraph darstellen, damit der Umbruch funktioniert
    cell_style = styles["Normal"].clone("cellstyle")
    cell_style.fontSize = 7.5
    cell_style.leading = 9

    header_style = styles["Normal"].clone("headerstyle")
    header_style.fontSize = 8
    header_style.textColor = colors.white
    header_style.fontName = "Helvetica-Bold"

    formatted_data = [[Paragraph(str(c), header_style) for c in data[0]]]
    for row in data[1:]:
        formatted_data.append(
            [Paragraph(str(row[0]), cell_style)] +
            [Paragraph(c.replace("\n", "<br/>"), cell_style) for c in row[1:]]
        )

    col_widths = [1.6 * cm] + [3.0 * cm] * len(categories)
    table = Table(formatted_data, colWidths=col_widths, repeatRows=1)

    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]
    table.setStyle(TableStyle(style))

    elements = [
        Paragraph("2. Board Heiztabelle (SP &amp; Budget kombiniert)", styles["Heading2"]),
        Spacer(1, 6),
        table,
    ]
    return elements


# ----------------------------------------------------------------------
# Tabelle 3: Board WP Netzbezug (Effizienz1-5, Effizienz0 wird ignoriert)
# ----------------------------------------------------------------------

def build_wp_netzbezug_table(wb, styles):
    if "Board WP Netzbezug" not in wb.sheetnames:
        return None
    ws = wb["Board WP Netzbezug"]
    header_map = get_header_map(ws)

    ws_col = col_letter_for(header_map, "WS")
    if ws_col is None:
        return None

    # Effizienz-Spalten finden (Effizienz 0 wird ausgeschlossen), diverse Schreibweisen zulassen
    eff_cols = []
    for name, col in header_map.items():
        m = re.fullmatch(r"Effizienz\s*(\d+)", name)
        if m:
            n = int(m.group(1))
            if n == 0:
                continue
            eff_cols.append((n, col))
    eff_cols.sort(key=lambda x: x[0])

    if not eff_cols:
        return None

    header_row = ["WS"] + [f"Effizienz {n}" for n, _ in eff_cols]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        v = row[ws_col - 1].value
        if not isinstance(v, (int, float)):
            continue
        line = [fmt_num(v)]
        for _, col in eff_cols:
            line.append(fmt_num(row[col - 1].value))
        rows.append((v, line))

    rows.sort(key=lambda x: x[0], reverse=True)  # 9 oben, 0 unten (bzw. absteigend)
    data = [header_row] + [r[1] for r in rows]

    if len(data) <= 1:
        return None

    col_widths = [1.6 * cm] + [2.3 * cm] * len(eff_cols)
    table = Table(data, colWidths=col_widths, repeatRows=1)

    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]
    table.setStyle(TableStyle(style))

    elements = [Paragraph("3. Board WP Netzbezug", styles["Heading2"]), Spacer(1, 6), table]
    return elements


# ----------------------------------------------------------------------
# Tabelle 4: Netzbezug Impact
# ----------------------------------------------------------------------

def build_netzbezug_impact_table(wb, styles):
    if "Netzbezug Impact" not in wb.sheetnames:
        return None
    ws = wb["Netzbezug Impact"]
    header_map = get_header_map(ws)

    wanted = ["Netzbezug", "Budget", "SP Runde 1", "SP Runde 2", "SP Runde 3", "SP Runde 4"]
    cols = [(name, header_map[name]) for name in wanted if name in header_map]
    if not cols or "Netzbezug" not in header_map:
        return None

    header_row = [name for name, _ in cols]
    data = [header_row]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        netzbezug_val = row[header_map["Netzbezug"] - 1].value
        if netzbezug_val is None:
            continue
        line = [fmt_num(row[col - 1].value) for _, col in cols]
        data.append(line)

    if len(data) <= 1:
        return None

    col_widths = [2.6 * cm] + [2.6 * cm] * (len(cols) - 1)
    table = Table(data, colWidths=col_widths, repeatRows=1)

    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]
    table.setStyle(TableStyle(style))

    elements = [Paragraph("4. Netzbezug Impact", styles["Heading2"]), Spacer(1, 6), table]
    return elements


# ----------------------------------------------------------------------
# Hauptprogramm
# ----------------------------------------------------------------------

def main():
    input_path = sys.argv[1] if len(sys.argv) > 1 else "Versionen/paper_draft_v1/game_data.xlsx"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "game_data_tables.pdf"

    wb = openpyxl.load_workbook(input_path, data_only=True)
    styles = getSampleStyleSheet()

    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    story = [Paragraph("Spielwerte-Tabellen", styles["Title"]), Spacer(1, 12)]

    builders = [
        build_base_values_table,
        build_heiztabelle_combined,
        build_wp_netzbezug_table,
        build_netzbezug_impact_table,
    ]

    first = True
    for builder in builders:
        elements = builder(wb, styles)
        if elements is None:
            continue
        if not first:
            story.append(PageBreak())
        story.extend(elements)
        first = False

    doc.build(story)
    print(f"PDF erstellt: {output_path}")


if __name__ == "__main__":
    main()